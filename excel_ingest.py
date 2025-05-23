import os
import json
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from copy import deepcopy
import asyncio
from datetime import datetime
import re

# Excel-specific imports
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlrd
import xlsxwriter

# LlamaIndex imports
from llama_index.core.schema import TextNode, MetadataMode
from llama_index.core import Settings

# Azure imports  
from azure.storage.blob.aio import BlobServiceClient
from azure.core.credentials import AzureKeyCredential

class ExcelProcessor:
    """Enhanced Excel file processor with support for complex structures."""
    
    def __init__(self, blob_connection_string: str, container_name: str):
        self.blob_connection_string = blob_connection_string
        self.container_name = container_name
        self.logger = logging.getLogger(__name__)
        
    async def process_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Main method to process Excel file and extract all content."""
        try:
            file_data = {
                "source_path": file_path,
                "file_name": Path(file_path).name,
                "sheets": [],
                "metadata": {},
                "images": [],
                "charts": []
            }
            
            # Load workbook with openpyxl for advanced features
            workbook = load_workbook(file_path, data_only=False)
            
            # Extract basic metadata
            file_data["metadata"] = await self._extract_file_metadata(workbook, file_path)
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                sheet_data = await self._process_worksheet(workbook[sheet_name], sheet_name)
                file_data["sheets"].append(sheet_data)
            
            # Extract images and charts
            file_data["images"] = await self._extract_images(workbook, file_path)
            file_data["charts"] = await self._extract_charts(workbook)
            
            workbook.close()
            return file_data
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file {file_path}: {str(e)}")
            raise

    async def _extract_file_metadata(self, workbook, file_path: str) -> Dict[str, Any]:
        """Extract file-level metadata."""
        metadata = {
            "file_name": Path(file_path).name,
            "file_size": os.path.getsize(file_path),
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "created_date": datetime.fromtimestamp(os.path.getctime(file_path)).isoformat(),
            "modified_date": datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
        }
        
        # Extract document properties if available
        props = workbook.properties
        if props:
            metadata.update({
                "title": props.title or "",
                "author": props.creator or "",
                "subject": props.subject or "",
                "description": props.description or "",
                "keywords": props.keywords or "",
                "company": getattr(props, 'company', '') or ""
            })
        
        return metadata

    async def _process_worksheet(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """Process individual worksheet with intelligent structure detection."""
        sheet_data = {
            "name": sheet_name,
            "tables": [],
            "summary": "",
            "structure_type": "unknown",
            "row_count": 0,
            "col_count": 0,
            "formulas": [],
            "named_ranges": []
        }
        
        # Get actual used range
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            # Empty sheet
            sheet_data["structure_type"] = "empty"
            return sheet_data
            
        # Convert to pandas for easier analysis
        data = []
        formulas = []
        
        for row in worksheet.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(str(cell.value))
                    # Capture formulas
                    if hasattr(cell, 'formula') and cell.formula:
                        formulas.append({
                            "cell": cell.coordinate,
                            "formula": cell.formula,
                            "value": cell.value
                        })
                else:
                    row_data.append("")
            data.append(row_data)
        
        if not data:
            sheet_data["structure_type"] = "empty"
            return sheet_data
            
        # Create DataFrame for analysis
        df = pd.DataFrame(data)
        df = df.dropna(how='all').dropna(axis=1, how='all')  # Remove empty rows/columns
        
        if df.empty:
            sheet_data["structure_type"] = "empty"
            return sheet_data
            
        sheet_data["row_count"] = len(df)
        sheet_data["col_count"] = len(df.columns)
        sheet_data["formulas"] = formulas
        
        # Detect structure type and extract tables
        tables = await self._detect_tables(df, sheet_name)
        sheet_data["tables"] = tables
        sheet_data["structure_type"] = await self._classify_sheet_structure(df, tables)
        
        # Generate summary
        sheet_data["summary"] = await self._generate_sheet_summary(df, tables, sheet_data)
        
        return sheet_data

    async def _detect_tables(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """Intelligently detect and extract tables from the worksheet."""
        tables = []
        
        # Strategy 1: Look for obvious table headers (first row has text, subsequent rows have data)
        if len(df) > 1:
            first_row = df.iloc[0]
            # Check if first row looks like headers (mostly text, not numbers)
            text_count = sum(1 for val in first_row if isinstance(val, str) and val.strip() and not val.replace('.', '').replace('-', '').isdigit())
            
            if text_count >= len(first_row) * 0.7:  # 70% text suggests headers
                table = await self._extract_table_with_headers(df, 0, sheet_name)
                if table:
                    tables.append(table)
            else:
                # Strategy 2: Look for patterns in the data
                tables.extend(await self._find_data_patterns(df, sheet_name))
        
        # Strategy 3: If no clear tables found, treat entire sheet as one table
        if not tables and not df.empty:
            tables.append(await self._create_generic_table(df, sheet_name))
        
        return tables

    async def _extract_table_with_headers(self, df: pd.DataFrame, header_row: int, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Extract table with identified headers."""
        if header_row >= len(df) - 1:
            return None
            
        headers = df.iloc[header_row].fillna("").astype(str).tolist()
        data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
        
        # Clean headers
        headers = [h.strip() if h.strip() else f"Column_{i+1}" for i, h in enumerate(headers)]
        
        # Create table structure
        table = {
            "name": f"{sheet_name}_table_{header_row + 1}",
            "headers": headers,
            "row_count": len(data_rows),
            "col_count": len(headers),
            "data_sample": data_rows.head(5).fillna("").astype(str).values.tolist(),
            "column_types": await self._analyze_column_types(data_rows),
            "summary_stats": await self._calculate_summary_stats(data_rows, headers),
            "text_content": await self._table_to_text(headers, data_rows)
        }
        
        return table

    async def _find_data_patterns(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """Find data patterns in sheets without clear headers."""
        tables = []
        
        # Look for sections separated by empty rows
        empty_row_indices = []
        for i, row in df.iterrows():
            if row.fillna("").astype(str).str.strip().eq("").all():
                empty_row_indices.append(i)
        
        # Split data by empty rows
        sections = []
        start_idx = 0
        for empty_idx in empty_row_indices:
            if empty_idx > start_idx:
                sections.append((start_idx, empty_idx))
            start_idx = empty_idx + 1
        
        if start_idx < len(df):
            sections.append((start_idx, len(df)))
        
        # Process each section
        for i, (start, end) in enumerate(sections):
            section_df = df.iloc[start:end].reset_index(drop=True)
            if not section_df.empty and len(section_df) > 1:
                table = await self._create_generic_table(section_df, f"{sheet_name}_section_{i+1}")
                tables.append(table)
        
        return tables

    async def _create_generic_table(self, df: pd.DataFrame, table_name: str) -> Dict[str, Any]:
        """Create a generic table structure for unstructured data."""
        # Try to infer if first row could be headers
        first_row = df.iloc[0] if len(df) > 0 else pd.Series()
        headers = [f"Column_{i+1}" for i in range(len(df.columns))]
        
        # Check if first row looks like headers
        if len(df) > 1:
            first_row_text_ratio = sum(1 for val in first_row if isinstance(val, str) and not str(val).replace('.', '').replace('-', '').isdigit()) / len(first_row)
            if first_row_text_ratio > 0.5:
                headers = first_row.fillna("").astype(str).tolist()
                headers = [h.strip() if h.strip() else f"Column_{i+1}" for i, h in enumerate(headers)]
                data_df = df.iloc[1:].reset_index(drop=True)
            else:
                data_df = df
        else:
            data_df = df
        
        return {
            "name": table_name,
            "headers": headers,
            "row_count": len(data_df),
            "col_count": len(headers),
            "data_sample": data_df.head(5).fillna("").astype(str).values.tolist(),
            "column_types": await self._analyze_column_types(data_df),
            "summary_stats": await self._calculate_summary_stats(data_df, headers),
            "text_content": await self._table_to_text(headers, data_df)
        }

    async def _analyze_column_types(self, df: pd.DataFrame) -> List[Dict[str, str]]:
        """Analyze data types for each column."""
        column_types = []
        
        for col_idx in range(len(df.columns)):
            col_data = df.iloc[:, col_idx].dropna()
            
            if col_data.empty:
                column_types.append({"type": "empty", "confidence": 1.0})
                continue
            
            # Convert to string for analysis
            str_data = col_data.astype(str).str.strip()
            
            # Check for different data types
            numeric_count = sum(1 for val in str_data if self._is_numeric(val))
            date_count = sum(1 for val in str_data if self._is_date(val))
            email_count = sum(1 for val in str_data if self._is_email(val))
            url_count = sum(1 for val in str_data if self._is_url(val))
            
            total_count = len(str_data)
            
            # Determine primary type
            if numeric_count / total_count > 0.8:
                col_type = "numeric"
                confidence = numeric_count / total_count
            elif date_count / total_count > 0.8:
                col_type = "date"
                confidence = date_count / total_count
            elif email_count / total_count > 0.8:
                col_type = "email"
                confidence = email_count / total_count
            elif url_count / total_count > 0.8:
                col_type = "url"
                confidence = url_count / total_count
            else:
                col_type = "text"
                confidence = 1.0 - max(numeric_count, date_count, email_count, url_count) / total_count
            
            column_types.append({"type": col_type, "confidence": confidence})
        
        return column_types

    def _is_numeric(self, value: str) -> bool:
        """Check if value is numeric."""
        try:
            float(value.replace(',', '').replace('$', '').replace('%', ''))
            return True
        except (ValueError, AttributeError):
            return False

    def _is_date(self, value: str) -> bool:
        """Check if value looks like a date."""
        date_patterns = [
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'\d{4}-\d{2}-\d{2}',
            r'\d{1,2}-\d{1,2}-\d{4}',
            r'\d{2}/\d{2}/\d{2}',
        ]
        return any(re.match(pattern, str(value)) for pattern in date_patterns)

    def _is_email(self, value: str) -> bool:
        """Check if value looks like an email."""
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        return bool(re.match(email_pattern, str(value)))

    def _is_url(self, value: str) -> bool:
        """Check if value looks like a URL."""
        url_pattern = r'https?://[^\s]+'
        return bool(re.match(url_pattern, str(value)))

    async def _calculate_summary_stats(self, df: pd.DataFrame, headers: List[str]) -> Dict[str, Any]:
        """Calculate summary statistics for the table."""
        stats = {
            "row_count": len(df),
            "column_count": len(headers),
            "non_empty_cells": int(df.notna().sum().sum()),
            "empty_cells": int(df.isna().sum().sum())
        }
        
        # Calculate per-column stats
        column_stats = []
        for i, header in enumerate(headers):
            if i < len(df.columns):
                col = df.iloc[:, i]
                col_stats = {
                    "column_name": header,
                    "non_empty_count": int(col.notna().sum()),
                    "unique_count": int(col.nunique()),
                    "most_common": str(col.mode().iloc[0]) if not col.mode().empty else ""
                }
                
                # If numeric, add numeric stats
                numeric_values = pd.to_numeric(col, errors='coerce').dropna()
                if not numeric_values.empty:
                    col_stats.update({
                        "mean": float(numeric_values.mean()),
                        "median": float(numeric_values.median()),
                        "min": float(numeric_values.min()),
                        "max": float(numeric_values.max())
                    })
                
                column_stats.append(col_stats)
        
        stats["columns"] = column_stats
        return stats

    async def _table_to_text(self, headers: List[str], df: pd.DataFrame) -> str:
        """Convert table to readable text format."""
        text_parts = []
        
        # Add table header
        text_parts.append(f"Table with {len(df)} rows and {len(headers)} columns:")
        text_parts.append("Columns: " + ", ".join(headers))
        text_parts.append("")
        
        # Add sample data
        text_parts.append("Sample data:")
        for i, row in df.head(10).iterrows():
            row_text = " | ".join([f"{header}: {str(row.iloc[j]) if j < len(row) else ''}" 
                                 for j, header in enumerate(headers)])
            text_parts.append(row_text)
        
        if len(df) > 10:
            text_parts.append(f"... and {len(df) - 10} more rows")
        
        return "\n".join(text_parts)

    async def _classify_sheet_structure(self, df: pd.DataFrame, tables: List[Dict]) -> str:
        """Classify the type of sheet structure."""
        if df.empty:
            return "empty"
        elif len(tables) == 1 and tables[0]["row_count"] > 10:
            return "data_table"
        elif len(tables) > 1:
            return "multi_table"
        elif df.shape[1] <= 3 and df.shape[0] <= 20:
            return "summary_report"
        else:
            return "mixed_content"

    async def _generate_sheet_summary(self, df: pd.DataFrame, tables: List[Dict], sheet_data: Dict) -> str:
        """Generate a natural language summary of the sheet."""
        summary_parts = []
        
        summary_parts.append(f"Sheet '{sheet_data['name']}' contains {sheet_data['row_count']} rows and {sheet_data['col_count']} columns.")
        
        if sheet_data["structure_type"] == "data_table":
            summary_parts.append("This appears to be a structured data table.")
        elif sheet_data["structure_type"] == "multi_table":
            summary_parts.append(f"This sheet contains {len(tables)} separate tables or data sections.")
        elif sheet_data["structure_type"] == "summary_report":
            summary_parts.append("This appears to be a summary or report format.")
        
        # Add information about tables
        for table in tables:
            summary_parts.append(f"Table '{table['name']}' has columns: {', '.join(table['headers'][:5])}")
            if len(table['headers']) > 5:
                summary_parts.append(f"... and {len(table['headers']) - 5} more columns.")
        
        # Add formula information
        if sheet_data.get("formulas"):
            summary_parts.append(f"The sheet contains {len(sheet_data['formulas'])} formulas.")
        
        return " ".join(summary_parts)

    async def _extract_images(self, workbook, file_path: str) -> List[Dict[str, Any]]:
        """Extract images from Excel file."""
        images = []
        
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Check for embedded images
                if hasattr(worksheet, '_images'):
                    for img in worksheet._images:
                        if isinstance(img, OpenpyxlImage):
                            # Extract and save image
                            image_data = {
                                "sheet": sheet_name,
                                "anchor": str(img.anchor) if hasattr(img, 'anchor') else "",
                                "width": img.width if hasattr(img, 'width') else 0,
                                "height": img.height if hasattr(img, 'height') else 0,
                            }
                            
                            # Save image file
                            image_filename = f"{Path(file_path).stem}_{sheet_name}_img_{len(images)}.png"
                            image_path = Path("temp_images") / image_filename
                            image_path.parent.mkdir(exist_ok=True)
                            
                            try:
                                img.save(str(image_path))
                                image_data["local_path"] = str(image_path)
                                image_data["filename"] = image_filename
                                images.append(image_data)
                            except Exception as e:
                                self.logger.warning(f"Could not save image: {str(e)}")
        
        except Exception as e:
            self.logger.warning(f"Error extracting images: {str(e)}")
        
        return images

    async def _extract_charts(self, workbook) -> List[Dict[str, Any]]:
        """Extract chart information from Excel file with safer imports."""
        charts = []
        
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Check for charts
                if hasattr(worksheet, '_charts'):
                    for chart in worksheet._charts:
                        # More generic approach without importing specific Chart class
                        chart_data = {
                            "sheet": sheet_name,
                            "type": chart.__class__.__name__,
                            "anchor": str(chart.anchor) if hasattr(chart, 'anchor') else "",
                        }
                        
                        # Try to extract title safely
                        try:
                            if hasattr(chart, 'title') and chart.title:
                                # Handle different title structures
                                if hasattr(chart.title, 'tx') and chart.title.tx:
                                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                                        chart_data["title"] = "Chart with Title"
                                    else:
                                        chart_data["title"] = str(chart.title.tx) if chart.title.tx else "Untitled Chart"
                                else:
                                    chart_data["title"] = "Untitled Chart"
                            else:
                                chart_data["title"] = "Untitled Chart"
                        except Exception:
                            chart_data["title"] = "Untitled Chart"
                        
                        charts.append(chart_data)
        
        except Exception as e:
            self.logger.warning(f"Error extracting charts: {str(e)}")
        
        return charts

    async def upload_images_to_blob(self, images: List[Dict[str, Any]]) -> Dict[str, str]:
        """Upload extracted images to Azure Blob Storage."""
        image_urls = {}
        
        if not images:
            return image_urls
        
        try:
            async with BlobServiceClient.from_connection_string(self.blob_connection_string) as blob_service_client:
                container_client = blob_service_client.get_container_client(self.container_name)
                
                # Ensure container exists
                try:
                    await container_client.create_container()
                except Exception:
                    pass  # Container might already exist
                
                for image in images:
                    if "local_path" in image and "filename" in image:
                        try:
                            blob_client = container_client.get_blob_client(image["filename"])
                            
                            with open(image["local_path"], "rb") as data:
                                await blob_client.upload_blob(data, overwrite=True)
                                image_urls[image["filename"]] = blob_client.url
                                
                        except Exception as e:
                            self.logger.error(f"Failed to upload image {image['filename']}: {str(e)}")
        
        except Exception as e:
            self.logger.error(f"Error uploading images to blob: {str(e)}")
        
        return image_urls

    def create_search_nodes(self, excel_data: Dict[str, Any], image_urls: Dict[str, str] = None) -> List[TextNode]:
        """Create LlamaIndex TextNodes from Excel data."""
        nodes = []
        
        if image_urls is None:
            image_urls = {}
        
        # Create a node for the overall file
        file_summary = self._create_file_summary(excel_data)
        file_node = TextNode(
            text=file_summary,
            metadata={
                "doc_id": Path(excel_data["source_path"]).stem,
                "content_type": "excel_file",
                "file_name": excel_data["file_name"],
                "sheet_count": len(excel_data["sheets"]),
                "node_type": "file_summary"
            }
        )
        nodes.append(file_node)
        
        # Create nodes for each sheet
        for sheet_idx, sheet in enumerate(excel_data["sheets"]):
            sheet_node = TextNode(
                text=sheet["summary"],
                metadata={
                    "doc_id": Path(excel_data["source_path"]).stem,
                    "content_type": "excel_sheet",
                    "sheet_name": sheet["name"],
                    "sheet_index": sheet_idx + 1,
                    "structure_type": sheet["structure_type"],
                    "node_type": "sheet_summary"
                }
            )
            nodes.append(sheet_node)
            
            # Create nodes for each table in the sheet
            for table_idx, table in enumerate(sheet["tables"]):
                table_node = TextNode(
                    text=table["text_content"],
                    metadata={
                        "doc_id": Path(excel_data["source_path"]).stem,
                        "content_type": "excel_table",
                        "sheet_name": sheet["name"],
                        "sheet_index": sheet_idx + 1,
                        "table_name": table["name"],
                        "table_index": table_idx + 1,
                        "row_count": table["row_count"],
                        "col_count": table["col_count"],
                        "headers": json.dumps(table["headers"]),
                        "node_type": "table_data"
                    }
                )
                nodes.append(table_node)
        
        # Create nodes for images if any
        for image in excel_data.get("images", []):
            if image["filename"] in image_urls:
                image_node = TextNode(
                    text=f"Image from sheet '{image['sheet']}' in Excel file. Image dimensions: {image.get('width', 'unknown')}x{image.get('height', 'unknown')} pixels.",
                    metadata={
                        "doc_id": Path(excel_data["source_path"]).stem,
                        "content_type": "excel_image",
                        "sheet_name": image["sheet"],
                        "image_path": image_urls[image["filename"]],
                        "image_filename": image["filename"],
                        "node_type": "image"
                    }
                )
                nodes.append(image_node)
        
        return nodes

    def _create_file_summary(self, excel_data: Dict[str, Any]) -> str:
        """Create a comprehensive summary of the Excel file."""
        summary_parts = []
        
        metadata = excel_data["metadata"]
        summary_parts.append(f"Excel file: {metadata['file_name']}")
        
        if metadata.get("title"):
            summary_parts.append(f"Title: {metadata['title']}")
        
        if metadata.get("author"):
            summary_parts.append(f"Author: {metadata['author']}")
        
        summary_parts.append(f"Contains {metadata['sheet_count']} sheets: {', '.join(metadata['sheet_names'])}")
        
        # Add sheet summaries
        for sheet in excel_data["sheets"]:
            if sheet["structure_type"] != "empty":
                summary_parts.append(f"Sheet '{sheet['name']}': {sheet['structure_type']} with {sheet['row_count']} rows and {sheet['col_count']} columns")
        
        # Add image and chart info
        if excel_data.get("images"):
            summary_parts.append(f"Contains {len(excel_data['images'])} images")
        
        if excel_data.get("charts"):
            summary_parts.append(f"Contains {len(excel_data['charts'])} charts")
        
        return ". ".join(summary_parts) + "."